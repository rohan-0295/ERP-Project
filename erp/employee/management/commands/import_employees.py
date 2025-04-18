from django.core.management.base import BaseCommand
from employee.models import Employee, EmployeeProfile
from department.models import Department
import openpyxl
import os

class Command(BaseCommand):
    help = 'Import employee data from Excel file'

    def handle(self, *args, **kwargs):
        try:
            file_path = os.path.join('Data', 'ERP_Large_Sample.xlsx')
            wb = openpyxl.load_workbook(file_path)
            sheet_name = 'Employees'
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"))
                return
            
            ws = wb[sheet_name]

            # Delete existing records
            EmployeeProfile.objects.all().delete()
            Employee.objects.all().delete()
            self.stdout.write(self.style.SUCCESS("Old employee and profile records deleted."))

            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    (
                        _id, first_name, last_name, email, phone,
                        _address, date_hired, dept_id, _emp_id,
                        job_title, employment_type, status
                    ) = row[:12]

                    department = Department.objects.filter(dept_id=dept_id).first()
                    if not department:
                        self.stdout.write(self.style.WARNING(f"Department '{dept_id}' not found. Skipping {first_name} {last_name}."))
                        continue

                    # Create employee
                    emp = Employee.objects.create(
                        job_title=job_title,
                        department=department,
                        employment_type=employment_type,
                        status=status,
                        date_hired=date_hired
                    )

                    # Create profile
                    EmployeeProfile.objects.create(
                        employee=emp,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone=str(phone),
                        gender='M'  # default gender if not in sheet
                    )

                    self.stdout.write(self.style.SUCCESS(f"Imported: {emp.employee_id} - {emp.job_title}"))

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error: {e}"))
                    continue

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Fatal error: {e}"))
