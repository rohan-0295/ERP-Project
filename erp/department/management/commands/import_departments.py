import os
import openpyxl
from django.core.management.base import BaseCommand
from department.models import Department
from django.conf import settings

class Command(BaseCommand):
    help = 'Import departments from Excel file'

    def handle(self, *args, **kwargs):
        file_path = os.path.join(settings.BASE_DIR, 'Data', 'ERP_Large_Sample.xlsx')

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_name = 'Departments'
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"Worksheet named '{sheet_name}' not found"))
                return

            ws = wb[sheet_name]
            self.stdout.write("\n>>\nOld department records deleted.")
            Department.objects.all().delete()

            for row in ws.iter_rows(min_row=2, values_only=True):
                dept_id, name, location = row[:3]
                if not (dept_id and name and location):
                    continue  # Skip rows with missing fields

                Department.objects.create(
                    dept_id=dept_id.strip(),
                    name=name.strip(),
                    location=location.strip()
                )
                self.stdout.write(self.style.SUCCESS(f"Imported: {dept_id.strip()} - {name.strip()}"))

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR("Excel file not found."))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {str(e)}"))
