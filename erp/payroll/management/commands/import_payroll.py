import os
import openpyxl
from django.core.management.base import BaseCommand
from payroll.models import Payroll
from employee.models import Employee  # Ensure you import the Employee model
from django.conf import settings

class Command(BaseCommand):
    help = 'Import payroll data from Excel file'

    def handle(self, *args, **kwargs):
        file_path = os.path.join(settings.BASE_DIR, 'Data', 'ERP_Large_Sample.xlsx')

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_name = 'Payroll'  # Change to the sheet where payroll data is located
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"Worksheet named '{sheet_name}' not found"))
                return

            ws = wb[sheet_name]
            self.stdout.write("\n>>\nOld payroll records deleted.")
            Payroll.objects.all().delete()  # Deletes all existing payroll records

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Read data from the Excel row
                payroll_employee_id = row[0]
                gross_salary = row[1]
                net_salary = row[2]
                payment_date = row[3]
                payment_method = row[4]
                
                try:
                    # Directly fetch Employee using the payroll_employee_id
                    employee = Employee.objects.get(id=payroll_employee_id)
                    
                    # Create new Payroll record
                    Payroll.objects.create(
                        employee=employee,
                        gross_salary=gross_salary,
                        net_salary=net_salary,
                        payment_date=payment_date,
                        payment_method=payment_method
                    )
                    self.stdout.write(self.style.SUCCESS(f"Payroll record imported for employee {payroll_employee_id}"))
                except Employee.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f"Employee with ID {payroll_employee_id} not found"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {str(e)}"))
